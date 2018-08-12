import click
from openpyxl import load_workbook
from copy import copy
import openpyxl
from openpyxl import Workbook

@click.command(name='my-cmd', context_settings=dict(
    ignore_unknown_options=True,
    allow_extra_args=True,
))
@click.option('--capitalize/--no-capitalize', default=False,required=False,
              help="To capitalize data or not")
@click.option('--preservestyles/--no-preservestyles', default=False,required=False,
              help="To preserve styles or not")

@click.argument('arguments', nargs=2)
def cli(capitalize,preservestyles,arguments):
	
    if capitalize:
        click.echo("Capitaliaze enabled")
        #ctx.obj1=True
    if preservestyles:
        click.echo("Preservestyles enabled")
        #ctx.obj2=True

    click.echo(arguments[0])
    click.echo(arguments[1])
	
    wb = openpyxl.load_workbook(filename=arguments[0])
    dest=openpyxl.Workbook()

    for w in wb.sheetnames:
        print(w)
        dest.create_sheet(title=w)

    for ws in wb.worksheets:
        
        w=dest[ws.title]
        for row,rowindex in enumerate(ws.rows, start=1):
        
            for cols,colindex in enumerate(ws.columns,start=1):
        
                if capitalize and preservestyles:
                    (w.cell(row=row,column=cols)).style=(ws.cell(row=row,column=cols)).style
                    w.cell(row=row,column=cols).value=(ws.cell(row=row,column=cols).value).capitalize()
        
                elif  preservestyles:
                    (w.cell(row=row,column=cols)).value=(ws.cell(row=row,column=cols).value)				
                    (w.cell(row=row,column=cols)).style=(ws.cell(row=row,column=cols)).style
                elif  capitalize:
                    (w.cell(row=row,column=cols)).value=(ws.cell(row=row,column=cols).value).capitalize()				
                    
                else:
                    w.cell(row=row,column=cols).value=(ws.cell(row=row,column=cols).value)
            
    dest.save(arguments[1])					
	


"""
import openpyxl
wb = openpyxl.load_workbook(filename='students.xlsx')

for ws in wb.worksheets:
    for row,rowindex in enumerate(ws.rows, start=1):
        #print (row)
        for cols,colindex in enumerate(ws.columns,start=1):
        #ws.cell(row=rowindex, column=colindex).value = (ws.cell(row=index, column=2).value).capitalize()
            ws.cell(row=row,column=cols).value=(ws.cell(row=row,column=cols).value).capitalize()

wb.save("capitalized.xlsx")

"""
            
""" 
			if cols.has_style:
                w.font = copy(ws.cell.font)
                w.border = copy(ws.cell.border)
                w.fill = copy(ws.cell.fill)
                w.number_format = copy(ws.cell.number_format)
                w.protection = copy(ws.cell.protection)
                w.alignment = copy(ws.cell.alignment)
				"""
"""	
	dest.save("capitalized1.xlsx")	
"""	
    #click.echo(arguments[0] + " " + arguments[1])
    
	
if __name__ == '__main__':
    		
	cli();