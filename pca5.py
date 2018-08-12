import sys
import click
import openpyxl as xl
import MySQLdb as sqlDB
from openpyxl import load_workbook

@click.group()
def sqldb():
    """Performs different commands on MySQL using python."""
    pass


def connect():
    """Connects to MySQL server."""
    try:
        db = sqlDB.connect(host="localhost", user="root", passwd="shanu")
        return db
    except MySQLdb.Error:
        print("Error occured while creating schema")
        return None;
    except MySQLdb.DatabaseError:
        print("Data base Error occured while creating schema")
        return None;
    except MySQLdb.IntegrityError:
        print("Integrity Error occured while creating schema")
        return None;
    except MySQLdb.NotSupportedError:
        print("not supported Error occured while creating schema")
        return None; 
    print("Connected to database")


def connect_db(dbname):
    """Connects to the given database of MySQL server."""
    try:
        db = sqlDB.connect(host="localhost", user="root", passwd="shanu", db=dbname)
        return db
    except MySQLdb.Error:
        print("Error occured while creating schema")
        return None;
    except MySQLdb.DatabaseError:
        print("Data base Error occured while creating schema")
        return None;
    except MySQLdb.IntegrityError:
        print("Integrity Error occured while creating schema")
        return None;
    except MySQLdb.NotSupportedError:
        print("not supported Error occured while creating schema")
        return None; 
    print("Connected to database"+dbname)
        


@sqldb.command('createdb', short_help="Creates a database and appropriate tables.")
@click.argument('dbname', nargs=1)
def createdb(dbname):
    """Creates a database."""
    
    db = connect()

    if db is None:
        click.echo("Error occurred")
        sys.exit(1)

    click.echo("Creating a database ...")
    
    sql_query = db.cursor()
    sql_cmd = f'CREATE DATABASE IF NOT EXISTS {dbname}'
    sql_query.execute(sql_cmd)
    sql_query.execute(f'USE {dbname}')
    db.commit()
    #db.close()

    #db = connect_db(dbname)
    sql_query = db.cursor()

    click.echo("Database is created with the name %s" % dbname)
    click.echo("Creating a table ...")

    table = "STUDENTS"

    db_type = """
                name varchar (255) not null,
                college varchar (10) not null,
                email_id varchar (255) not null,
                dbname varchar (255) not null primary key
        """
    sql_query.execute(f"CREATE TABLE IF NOT EXISTS {table} ({db_type})")
    db.commit()

    click.echo("Table is created with the name %s" % table)
    click.echo("Creating a table ...")

    table = "MARKS"
    db_type = """
                name varchar(255) not null primary key,
                clg varchar(20) not null,
                test1 int not null,
                test2 int not null,
                test3 int not null,
                test4 int not null,
                total int not null
        """
    sql_query.execute(f'CREATE TABLE IF NOT EXISTS {table} ({db_type})')
    db.commit()

    click.echo("Table is created with the name %s" % table)
	
    click.echo("Adding foreign key ...")

    db_type = """
                   add constraint students_name
                   FOREIGN KEY (name)
                   REFERENCES students(dbname)
                   ON DELETE NO ACTION
                   ON UPDATE NO ACTION
           """

    sql_query.execute(f"ALTER TABLE {table} {db_type};")

    click.echo("Foreign Key is successfully added to the table %s" % table)
    
    table="COLLEGES"
    
    db_type = """
                collegename varchar(255) not null,
                acronym varchar(10) not null primary key,
                location varchar(100) not null,
                contact varchar(50) not null
        """
    sql_query.execute(f'CREATE TABLE IF NOT EXISTS {table} ({db_type})')
    db.commit()

    click.echo("Table is created with the name %s" % table)
	
    click.echo("Adding foreign key ...")

    db_type = """
                   add constraint college_name
                   FOREIGN KEY (acronym)
                   REFERENCES students(college)
                   ON DELETE NO ACTION
                   ON UPDATE NO ACTION
           """

    #sql_query.execute(f"ALTER TABLE {table} {db_type};")

    click.echo("Foreign Key is successfully added to the table %s" % table)
    
    db.commit()
    db.close()


@sqldb.command('dropdb', short_help="Drops the database.")
@click.argument('dbname', nargs=1)
def dropdb(dbname):
    """Drops the database."""

    db = connect()
    if db is None:
        click.echo("Database not found ....exiting")
        sys.exit(1)

    click.echo("Dropping a database ...")

    sql_query = db.cursor()
    sql_cmd = f'DROP DATABASE IF EXISTS {dbname}'
    sql_query.execute(sql_cmd)
    db.commit()
    db.close()

    click.echo("Database %s is dropped" % dbname)


@sqldb.command('import_data', short_help="Imports data into tables")
@click.argument('dbname', nargs=1)
def import_data(dbname):
    """Imports data from input into appropriate tables."""
    db = connect_db(dbname)
    if db is None:
        click.echo("database not found...")
        sys.exit(1)

    click.echo("Importing data ...")

    sql_query = db.cursor()

    wb = xl.load_workbook("students.xlsx")
    ws = wb[wb.sheetnames[0]]

    table = "STUDENTS"

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value.lower()
        college = ws.cell(row=row, column=2).value.lower()
        email_id = ws.cell(row=row, column=3).value.lower()
        dbname = ws.cell(row=row, column=4).value.lower()
        sql_query.execute(f"INSERT INTO {table} VALUES ('{name}', '{college}', '{email_id}', '{dbname}')")
        db.commit()

    click.echo("Data imported to the table %s" % table)

    
    
    worksheet = wb['Colleges']
    query = "INSERT INTO colleges (collegename,acronym,location,contact) VALUES (%s,%s,%s,%s)"
    
    cursor = db.cursor()
   
    for row in worksheet.iter_rows():
        s1 = row[0].value
        s2 = row[1].value
        s3 = row[2].value
        s4 = row[3].value
        
        cursor.execute(query, (s1, s2, s3, s4))
            
            
    wb = xl.load_workbook("finall.xlsx")
    ws = wb["Mock Test Summary for 2016 PP "]
    
    table = "MARKS"
    query = "INSERT INTO marks (name,clg,test1,test2,test3,test4,total) VALUES (%s,%s,%s,%s,%s,%s,%s)"
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        value_list=name.split("_")
        n=value_list[2].lower()
        clg=value_list[1].lower()	
        test1 = int(ws.cell(row=row, column=2).value)
        test2 = int(ws.cell(row=row, column=3).value)
        test3 = int(ws.cell(row=row, column=4).value)
        test4 = int(ws.cell(row=row, column=5).value)
        total = int(ws.cell(row=row, column=6).value)
        
        q='SELECT COUNT(*) FROM colleges WHERE acronym=%s'
        ans=0
        cursor.execute(q,(clg,))
        ans=cursor.fetchall()
        print(ans)
        if(ans[0][0]!=0):
            cursor.execute(query, (n,clg,test1,test2,test3,test4,total))
        
        
        #sql_query.execute(f"INSERT INTO {table} VALUES ('{name}', {test1}, {test2}, {test3}, {test4}, {total})")
        db.commit()

    click.echo("Data imported to the table %s" % table)

    db.commit()
    db.close()


@sqldb.command('college_stats', short_help="Prints out a console report showing College acronym, count, min, avg, max")
@click.argument('dbname', nargs=1)
def college_stats(dbname):
    """Print out a console report showing number of students per college and the minimum, maximum and average marks of
       students from the college in the form
       [College Acronym, Student Count, Minimum marks, Average marks, Maximum marks]
    """
    db = connect_db(dbname)
    if db is None:
        click.echo("database not found....")
        sys.exit(1)

    click.echo("Generating college stats ...")
    sql_query = db.cursor()
    
    cursor = db.cursor()
    sql_cmd = """
            students s INNER JOIN marks m
            ON s.dbname = m.name
    """
    #sql_query.execute(f"""SELECT colleges.collegename, count(*), MIN(m.total), MAX(m.total), AVG(m.total) FROM {sql_cmd}
    #                  GROUP BY s.college""")
    
    query='SELECT COLLEGES.collegename,COUNT(*),MIN(total),MAX(total),AVG(total) FROM marks,COLLEGES where COLLEGES.acronym=marks.clg GROUP BY marks.clg'
    cursor.execute(query)
    
    click.echo("College stats are generated")

    stats = cursor.fetchall()

    click.echo("College Acronym | Student Count | Minimum Marks | Maximum Marks | Average Marks")

    for data in stats:
        click.echo(data[0].center(55) + " | " +
                   str(data[1]).center(13) + " | " +
                   str(data[2]).center(13) + " | " +
                   str(data[3]).center(13) + " | " +
                   str(data[4]).strip().center(13))


if __name__ == '__main__':
    sqldb()
