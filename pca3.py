import click
import openpyxl
from openpyxl import Workbook
import os
import bs4
from bs4 import BeautifulSoup
import urllib.request
from openpyxl.styles import Color, Fill,Font


@click.command()
@click.argument('inputurl',nargs=1)
@click.argument('outputxlsx',nargs=1)
def parser(inputurl,outputxlsx):

    """output data"""
    wb = openpyxl.Workbook()
    sheet = wb.active


    """ opening browser"""
    source = urllib.request.urlopen('https://d1b10bmlvqabco.cloudfront.net/attach/inpg92dp42z2zo/hdff4poirlh7i6/io5hun2sdr21/mock_results.html').read()
    soup=BeautifulSoup(source,'html.parser')

    
    i=soup.find_all('h1')[0]
    sheet.title=i.text.strip()[:30]
    
    """copying headings"""
    table=soup.find_all('table')[0]
    tr=table.find_all('tr')
    th=table.find_all('th')[1:]
   
    
    for index,i in enumerate(th,start=1):
        sheet.cell(row=1, column=index).value = i.text.strip()
        sheet.cell(row=1,column=index).font = Font(bold=True)
        
    """copying data"""

    table = soup.find_all('table')[0]


    for rows,i in enumerate(table.find_all('tr'),start=1):
        td=i.find_all('td')
        
        td=td[1:]
        for cols,j in enumerate(td,start=1):
            sheet.cell(row=rows, column=cols).value =j.text.strip()
            


    wb.save(outputxlsx)


if __name__=='__main__':
    parser()
